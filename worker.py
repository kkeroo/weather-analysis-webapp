import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment

months_names = ["Januar", "Februar", "April", "Maj", "Junij", "Julij", "Avgust", "September", "Oktober", "November", "December"]

def clear_data(df):

    dates = list(df["Date"])
    temps = list(df["Temperature"])

    wrong_data_num = 0

    for i in range(len(dates)):
        month = dates[i].month
        temperature = temps[i]

        if month >= 4 and month <= 10:
            if temperature < -10 or temperature > 55:
                temps[i] = np.nan
                wrong_data_num += 1
        else:
            if temperature < -20 or temperature > 35:
                temps[i] = np.nan
                wrong_data_num += 1
    
    df['Temperature'] = temps
    return df, wrong_data_num


def interpolate_temperature(temp):
    orig_len = len(temp)
    missing_data = len(temp[temp.str.contains("---")])
    temp[temp.str.contains("---")] = np.nan
    temp = pd.to_numeric(temp)
    temp = temp.interpolate(limit=415)

    left_missing = len(temp[temp.isnull()])
    interpolated = missing_data - left_missing

    print("<----ORIGINAL---->")
    print(f"Original Data length: {orig_len}, missing data: {missing_data}/{orig_len} \n \
        interpolated data: {interpolated}, left missing data: {left_missing}")
    print("<------------->")

    data_info = {'original': orig_len, 'missing': missing_data, 'interpolated': interpolated, 'left_missing': left_missing}

    return temp, data_info

def interpolate_daily_stats(daily):
    orig_len = len(daily)
    missing = len(daily[daily["Avg. temp"].isnull()])
    daily["Avg. temp"] = daily["Avg. temp"].interpolate(limit=2)
    daily["Temp std."] = daily["Temp std."].interpolate(limit=2)
    daily["Temp max"] = daily["Temp max"].interpolate(limit=2)
    daily["Temp min"] = daily["Temp min"].interpolate(limit=2)
    left_missing = len(daily[daily["Avg. temp"].isnull()])
    interpolated = missing - left_missing

    print("<----DAILY---->")
    print(f"Original Data length: {orig_len}, missing data: {missing}/{orig_len} \n \
        interpolated data: {interpolated}, left missing data: {left_missing}")
    print("<------------->")

    data_info = {'original': orig_len, 'missing': missing, 'interpolated': interpolated, 'left_missing': left_missing}

    return daily, data_info

def temperature_avg_std(df):
    temperature_cleaned = df.iloc[:, 3]
    # print(temperature_cleaned)
    # temperature_cleaned = temperature[~temperature.str.contains("---")]
    temperature_cleaned = pd.to_numeric(temperature_cleaned)
    temperature_mean = np.mean(temperature_cleaned)
    temp_std = np.std(temperature_cleaned)
    temp_max = np.max(temperature_cleaned)
    temp_min = np.min(temperature_cleaned)
    return round(temperature_mean, 2), round(temp_std, 2), round(temp_max, 2), round(temp_min, 2)

# izračuna mesečno količino dežja in poišče dan ko je padlo največ dežja
def rain_total(df):
    rain = df.iloc[:, 5]
    rain_cleaned = rain
    if (rain.dtype != float):
        rain_cleaned = rain[~rain.str.contains("---")]
        rain_cleaned = pd.to_numeric(rain_cleaned)
    rain_sum = np.sum(rain_cleaned)
    rain_max = np.max(rain_cleaned)
    # rain_max_day = df[df['Rain'] == rain_max]
    # rain_max_day = rain_max_day.iloc[0,0]
    try:
        rain_max_day = df[df['Rain'] == rain_max]
        rain_max_day = rain_max_day.iloc[0,0]
    except:
        rain_max_day = ""
    return round(rain_sum, 2), rain_max_day

def wind_dir(df):
    wind = df.iloc[:, 6]
    wind_cleaned = wind[~wind.str.contains("---")]
    wind_freq = wind_cleaned.value_counts()
    biggest_wind_freq = "/"
    try:
        biggest_wind_freq = wind_freq.index[0]
    except:
        biggest_wind_freq = "/"
    return biggest_wind_freq

# funkicja sprejme dnevne podatke v enem letu in
# vrne top dež za vsak mesec
def top_rainy_day_in_month(df):
    # df: daily df
    curr_month = df.iloc[0,0].month
    curr_year = df.iloc[0,0].year
    last_month = df.iloc[-1,0].month
    monthly = []

    top_rain_list = list()
    top_rain_day_list = list()

    while curr_month <= last_month:
        if curr_month == 12:
            monthly = df[(df['dt'] >= datetime(curr_year, curr_month, 1)) & (df['dt'] <= datetime(curr_year, curr_month, 31))]
        else:
            monthly = df[(df['dt'] >= datetime(curr_year, curr_month, 1)) & (df['dt'] < datetime(curr_year, curr_month+1, 1))]
        top_rain = np.max(monthly['Total rain'])
        #print(top_rain)
        top_rain_day = monthly[monthly['Total rain'] == top_rain]
        top_rain_day = datetime.date(top_rain_day.iloc[0,0])
        #print(top_rain_day)
        top_rain_list.append(top_rain)
        top_rain_day_list.append(top_rain_day)
        curr_month = curr_month + 1
    return top_rain_list, top_rain_day_list


def yearly_stats(df, start_month, end_month, daily_df):
    curr_year = df.iloc[0,0].year

    # najprej statistika za celo leto
    yearly_temp_avg, yearly_temp_std, year_temp_max, year_temp_min = temperature_avg_std(df)
    yearly_rain_total, yearly_rain_max = rain_total(df)
    yearly_wind_dir = wind_dir(df)
    yearly_df = pd.DataFrame([[curr_year, yearly_temp_avg, yearly_temp_std, yearly_rain_total, yearly_wind_dir]], columns=['Year',  'Avg. temp', 'Temp std.', 'Total rain', 'Wind dir.'])

    # potem pa statistika za vsak mesec v letu
    monthly_df = pd.DataFrame(columns=['Month', 'Avg. temp', 'Temp std.', 'Total rain', 'Max rain', 'Wind dir.'])
    months = list()
    avg_temps = list()
    temp_stds = list()
    temp_maxs = list()
    temp_mins = list()
    rains = list()
    #max_rains = list()
    wind_dirs = list()
    for m in range(start_month, end_month+1):
        monthly = []
        if m < 12:
            monthly = df[(df['dt'] >= datetime(curr_year, m, 1)) & (df['dt'] < datetime(curr_year, m+1, 1))]
        else:
            monthly = df[(df['dt'] >= datetime(curr_year, m, 1)) & (df['dt'] < datetime(curr_year+1, 1, 1))]

        monthly_temp_mean, monthly_temp_std, monthly_temp_max, monthly_temp_min = temperature_avg_std(monthly)
        monthly_rain, monthly_rain_max = rain_total(monthly)
        monthly_wind_dirs = wind_dir(monthly)
        months.append(m)
        avg_temps.append(monthly_temp_mean)
        temp_stds.append(monthly_temp_std)
        temp_maxs.append(monthly_temp_max)
        temp_mins.append(monthly_temp_min)
        rains.append(monthly_rain)
        #max_rains.append(monthly_rain_max)
        wind_dirs.append(monthly_wind_dirs)
        #print("Povprečna temperatura v mesecu {} je {}, std je {}, deža je padlo {}, največ dne {}".format(m, monthly_temp_mean, monthly_temp_std, monthly_rain, monthly_rain_max))
    max_rains, max_rain_days = top_rainy_day_in_month(daily_df)
    monthly_df['Month'] = months
    monthly_df['Avg. temp'] = avg_temps
    monthly_df['Temp std.'] = temp_stds
    monthly_df['Temp max'] = temp_maxs
    monthly_df['Temp min'] = temp_mins
    monthly_df['Total rain'] = rains
    monthly_df['Max rain'] = max_rains
    monthly_df['Max rain day'] = max_rain_days
    monthly_df['Wind dir.'] = wind_dirs
    #print(monthly_df)
    return yearly_df, monthly_df

def calculate_gdd(df, degree):
    tmaxlist = list(df['Temp max'])
    tminlist = list(df["Temp min"])
    gdd = list()
    for i in range(len(tmaxlist)):
        if not np.isnan(tmaxlist[i]) and not np.isnan(tminlist[i]):
            # tmax.append(tmaxlist[i], 30)
            M = min(max(tmaxlist[i], degree), 30)
            m = max(tminlist[i], degree)
            gdd.append((M+m) / 2 - degree)
        else:
            gdd.append(0)

    gdd = np.array(gdd)
    gdd_total = np.sum(gdd)
    return gdd, gdd_total


def daily_stats_in_year(df, degree):
    curr_date = df.iloc[0,0]
    last_date = df.iloc[-1,0]

    dts = list()
    days = list()
    avg_temps = list()
    temp_stds = list()
    temp_maxs = list()
    temp_mins = list()
    rains = list()
    wind_dirs = list()

    while curr_date <= last_date:
        daily_df = df[df['Date'] == datetime.date(curr_date)]
        daily_temp_avg, daily_temp_std, daily_temp_max, daily_temp_min = temperature_avg_std(daily_df)
        daily_rain, daily_max_rain = rain_total(daily_df)
        daily_wind_dirs = wind_dir(daily_df)
        dts.append(curr_date)
        days.append(datetime.date(curr_date))
        avg_temps.append(daily_temp_avg)
        temp_stds.append(daily_temp_std)
        temp_maxs.append(daily_temp_max)
        temp_mins.append(daily_temp_min)
        rains.append(daily_rain)
        wind_dirs.append(daily_wind_dirs)

        curr_date = curr_date + timedelta(days=1)
    
    daily_df = pd.DataFrame()
    daily_df['dt'] = dts
    daily_df['Day'] = days
    daily_df['Avg. temp'] = avg_temps
    daily_df['Temp std.'] = temp_stds
    daily_df['Temp max'] = temp_maxs
    daily_df['Temp min'] = temp_mins
    daily_df['Total rain'] = rains
    daily_df['Wind dir.'] = wind_dirs

    daily_df, data_info = interpolate_daily_stats(daily_df)

    gdd, gdd_total = calculate_gdd(daily_df, degree)
    daily_df['Gdd'] = gdd
    return daily_df, gdd_total, data_info


#data_type: monthly/daily
def write_data_to_ws(df, ws, year, data_type, gdd, degree, data_info):
    #set column width
    if (data_type == 'monthly'):
        ws.column_dimensions['A'].width = 9
    else:
        ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 11
    ws.column_dimensions['G'].width = 11
    if data_type == 'monthly':
        ws.column_dimensions['H'].width = 11
        ws.column_dimensions['I'].width = 15
    else:
        ws.column_dimensions['H'].width = 11

    day_or_month = 'Month'
    if data_type == 'daily':
        day_or_month = 'Day'
        ws.append([day_or_month, 'Avg. temperature', 'Temperature std.', 'Max temperature', 'Min temperature', 'Total rain', 'Wind dir.', 'GDD'])
    else:
        ws.append([day_or_month, 'Avg. temperature', 'Temperature std.', 'Max temperature', 'Min temperature', 'Total rain', 'Wind dir.', 'Most rain', 'Most rainy day'])
    for index, row in df.iterrows():
        if (data_type == 'monthly'):
            r = [row[day_or_month], row['Avg. temp'], row['Temp std.'], row['Temp max'], row['Temp min'], row['Total rain'], row['Wind dir.'], row['Max rain'], row['Max rain day']]
            ws.append(r)
        else:
            r = [row[day_or_month], row['Avg. temp'], row['Temp std.'], row['Temp max'], row['Temp min'], row['Total rain'], row['Wind dir.'], row['Gdd']]
            ws.append(r)
    
    num_rows = df.shape[0] + 1
    ref_column = 'H'
    if data_type == 'monthly':
        ref_column = 'I'
    tab = Table(displayName="{}_stats_{}".format(data_type, year), ref="A1:{}{}".format(ref_column, num_rows))
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    if (data_type == 'daily'):
        add_gdd_to_ws(ws, gdd, degree)
        add_daily_data_info(ws, data_info['original'], data_info['missing'], data_info['interpolated'], data_info['left_missing'])

    datum = 'Dan'
    if (data_type == 'monthly'):
        datum = 'Mesec'
    create_rain_chart(ws, num_rows+1, datum)
    create_temp_chart(ws, num_rows+1, datum)

def add_gdd_to_ws(ws, gdd, degree):
    ws['K2'] = "GDD ({})".format(degree)
    cell = ws['K2']
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.style = 'Accent1'
    cell = ws['K3']
    cell.value =gdd
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.style = '20 % - Accent1'

def add_daily_data_info(ws, original, missing, interpolated, left_missing):
    ws['M2'] = '# Days'
    ws['M3'] = '# Missing days'
    ws['M4'] = '# Interpolated days'
    ws['M5'] = '# Left-missing days'

    ws['M2'].style = 'Accent1'
    ws['M3'].style = 'Accent1'
    ws['M4'].style = 'Accent1'
    ws['M5'].style = 'Accent1'

    ws['N2'] = original
    ws['N3'] = missing
    ws['N4'] = interpolated
    ws['N5'] = left_missing

    ws['N2'].style = '20 % - Accent1'
    ws['N3'].style = '20 % - Accent1'
    ws['N4'].style = '20 % - Accent1'
    ws['N5'].style = '20 % - Accent1'

    ws.column_dimensions['M'].width = 18

def add_data_info(ws, original, missing, interpolated, left_missing, wrong):
    ws['J1'] = '# Measurements'
    ws['J2'] = '# Missing measurements'
    ws['J3'] = '# Interpolated measurements'
    ws['J4'] = '# Left-missing measurements'
    ws['J5'] = '# Wrong measurements'

    ws['J1'].style = 'Accent1'
    ws['J2'].style = 'Accent1'
    ws['J3'].style = 'Accent1'
    ws['J4'].style = 'Accent1'
    ws['J5'].style = 'Accent1'

    ws['K1'] = original
    ws['K2'] = missing
    ws['K3'] = interpolated
    ws['K4'] = left_missing
    ws['K5'] = wrong

    ws['K1'].style = '20 % - Accent1'
    ws['K2'].style = '20 % - Accent1'
    ws['K3'].style = '20 % - Accent1'
    ws['K4'].style = '20 % - Accent1'
    ws['K5'].style = '20 % - Accent1'

    ws.column_dimensions['J'].width = 25

def create_rain_chart(ws, num_of_rows, date_type):
    chart = BarChart()
    chart.type = "col"
    chart.style = 2
    chart.title = "Padavine"
    chart.y_axis.title = 'Padavine [mm]'
    chart.x_axis.title = date_type
    data = Reference(ws, min_col=6, min_row = 2, max_row = num_of_rows)
    labels = Reference(ws, min_col=1, min_row = 2, max_row = num_of_rows)
    chart.add_data(data)
    chart.set_categories(labels)
    chart.shape = 4
    chart.legend = None
    ws.add_chart(chart, "K7")

def create_temp_chart(ws, num_of_rows, date_type):
    chart = LineChart()
    chart.style = 2
    chart.title = "Povprečne temperature"
    chart.y_axis.title = 'Temperatura'
    chart.x_axis.title = date_type
    data = Reference(ws, min_col=2, min_row = 2, max_row = num_of_rows)
    labels = Reference(ws, min_col=1, min_row = 2, max_row = num_of_rows)
    chart.add_data(data)
    chart.set_categories(labels)
    chart.shape = 4
    chart.legend = None
    ws.add_chart(chart, "K23")   

def write_dataframe_to_ws(df,ws, data_info, wrong_data):
    ws.append(['Date', 'Time', 'Temperature', 'Humidity', 'Rain', 'Wind_dir', 'Wind_speed', 'Gust'])
    for index, row in df.iterrows():
        t = row['Temperature']
        try:
            t = pd.to_numeric(t)
        except:
            pass
        
        h = row['Humidity']
        try:
            h = pd.to_numeric(h)
        except:
            pass
        rain = row['Rain']
        try:
            rain = pd.to_numeric(rain)
        except:
            pass
        s = row['Wind_speed']
        try:
            s = pd.to_numeric(s)
        except:
            pass
        g = row['Gust']
        try:
            g = pd.to_numeric(g)
        except:
            pass
        r = [row['Date'], row['Time'], t, h, rain, row['Wind_dir'], s, g]
        ws.append(r)
    
    num_rows = df.shape[0] + 1
    tab = Table(displayName="dataframe", ref="A1:H{}".format(num_rows))
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    add_data_info(ws, data_info['original'], data_info['missing'], data_info['interpolated'], data_info['left_missing'], wrong_data)
    

def crop_date_range(dataframe, start_date, end_date):
    date_arr = start_date.split('/')
    start_day = int(date_arr[0])
    start_month = int(date_arr[1])
    start_year = int(date_arr[2])

    date_arr = end_date.split('/')
    end_day = int(date_arr[0])
    end_month = int(date_arr[1])
    end_year = int(date_arr[2])

    df = dataframe[(dataframe["dt"] >= datetime(start_year, start_month, start_day)) & (dataframe["dt"] < datetime(end_year, end_month, end_day))]

    return df

def generate(df, start_date, end_date, base_temp):
    dataframe = df
    # Timestamp
    timestamp_col = dataframe.iloc[:, 1]
    date_list = list()
    time_list = list()
    datetime_list = list()
    for timestamp_el in timestamp_col:
        date_obj = datetime.fromtimestamp(timestamp_el)
        datetime_list.append(date_obj)
        date_list.append(datetime.date(date_obj))
        time_list.append(datetime.time(date_obj))

    # Add Date and Time column to dataframe and rename columns
    # also drop the original date & time column and timestamp
    dataframe["Date"] = date_list
    dataframe["Time"] = time_list
    dataframe["dt"] = datetime_list
    dataframe = dataframe.drop(columns=dataframe.iloc[:, :2], axis=1)
    dataframe = dataframe.set_axis(['Temperature', 'Humidity', 'Rain', 'Wind_dir', 'Wind_speed', 'Gust', 'Fail', 'Date', 'Time', 'dt'], axis=1)
    dataframe = dataframe.drop(columns=['Fail'], axis=1)

    # reoorder columns
    dataframe = dataframe[['dt', 'Date', 'Time', 'Temperature', 'Humidity', 'Rain', 'Wind_dir', 'Wind_speed', 'Gust']]

    # keep data between starting and ending date
    dataframe = crop_date_range(dataframe, start_date, end_date)

    # clear data
    dataframe['Temperature'], data_info = interpolate_temperature(dataframe['Temperature'])
    dataframe, wrong_data_num = clear_data(dataframe)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Podatki'
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 7
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 7

    write_dataframe_to_ws(dataframe, ws, data_info, wrong_data_num)

    # Monthly statistics
    years = list()
    dates = dataframe["Date"]
    for dt in dates:
        y = dt.year
        if not y in years:
            years.append(y)

    for year in years:
        one_year = dataframe[(dataframe["dt"] >= datetime(year, 1, 1)) & (dataframe["dt"] < datetime(year, 12, 31))]
        daily_df, gdd, daily_data_info = daily_stats_in_year(one_year, base_temp)
        #gdd = calculate_gdd(daily_df, 10)
        first_month = one_year.iloc[0,0].month
        last_month = one_year.iloc[-1,0].month
        yearly_df, monthly_df = yearly_stats(one_year, first_month, last_month, daily_df)
        
        # ustvarimo mesečne podatke
        ws = wb.create_sheet('Meseci-{}'.format(year))
        write_data_to_ws(monthly_df, ws, year, 'monthly', gdd, base_temp, data_info=None)
        # ustvarimo dnevne podatke
        ws = wb.create_sheet('Dnevi-{}'.format(year))
        write_data_to_ws(daily_df, ws, year, 'daily', gdd, base_temp, data_info=daily_data_info)

    # wb.save(f'{filename}.xlsx')
    return 2

if __name__ == "__main__":
    generate('weatherstation_export.csv', '1/1/2021', '31/12/2021', 8)